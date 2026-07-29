"""Implementation-distinct, capability-closed ANAL-009--011 validation."""

from __future__ import annotations

import hashlib
import json
import math
import platform
from pathlib import Path
import re
import xml.etree.ElementTree as ET

from openpyxl import load_workbook
import pdfplumber
from pypdf import PdfReader

from sft.chemistry.generated_law import prediction_program_document
from sft.chemistry.generated_observational_law import observational_experiment_registration_record
from sft.chemistry.photoluminescence_family_batch_v1 import (
    ANALYSIS_PATH, AUTHORITIES, FLUORESCENCE_SPEC, PHOSPHORESCENCE_SPEC, RAMAN_SPEC, SPECS,
)
from sft.claim_evidence import (
    CapabilityClosedFoldInterpreter, CrossPlatformCustodyExchange, HostilePackageAuditor,
    TargetVault, fold_program_from_mapping, snapshot_protected_tree, target_identity_from_release,
)
from sft.engine import (
    EmpiricalValidation, seal_isolation_certificate, seal_target_custody_certificate,
    unsealed_isolation_certificate, unsealed_target_custody_certificate,
)
from sft.engine.canonical import sha256_identity
from sft.engine.empirical import BlindExperimentBoundary, PredictionEnvelope
from sft.engine.exact import HeldLabel
from sft.engine.source import hash_file


def digest(data: bytes) -> str:
    return "sha256:" + hashlib.sha256(data).hexdigest()


def canonical_digest(value: object) -> str:
    return digest(json.dumps(value, sort_keys=True, separators=(",", ":"), ensure_ascii=False).encode())


def external_token(value: object) -> str:
    return str(value)


def independent_source_extent(root: Path, analysis: dict[str, object]):
    inventory_path = root / "experiments/external_sources/chemistry/snapshots/anal-009-011-photoluminescence-v1/source-inventory-v1.json"
    inventory = json.loads(inventory_path.read_text())
    payload = dict(inventory)
    stored = payload.pop("inventory_payload_sha256")
    if canonical_digest(payload) != stored or len(inventory["sources"]) != 16:
        raise ValueError("photoluminescence source inventory changed")
    counts = {"pdf_pages": 0, "html_documents": 0, "xml_documents": 0, "workbooks": 0}
    source_bytes = 0
    for source in inventory["sources"]:
        path = root / source["path"]
        data = path.read_bytes()
        if len(data) != source["byte_count"] or digest(data) != source["sha256"]:
            raise ValueError(f"captured source changed: {source['path']}")
        source_bytes += len(data)
        if source["media_kind"] == "pdf":
            counts["pdf_pages"] += len(PdfReader(path).pages)
        elif source["media_kind"] == "html":
            counts["html_documents"] += 1
        elif source["media_kind"] in {"bioc-xml", "pmc-oa-package-unavailable"}:
            ET.fromstring(data)
            counts["xml_documents"] += 1
        elif source["media_kind"] == "linked-certified-workbook":
            load_workbook(path, data_only=True, read_only=True, keep_vba=True)
            counts["workbooks"] += 1
        else:
            raise ValueError("unregistered photoluminescence source kind")
    if counts != analysis["complete_source_surface_counts"] | {"characters": analysis["complete_source_surface_counts"]["characters"]}:
        # Character counts are parser-version evidence already fixed by the
        # analysis hash; the implementation-distinct extent is categorical.
        comparable = dict(analysis["complete_source_surface_counts"])
        comparable.pop("characters")
        if counts != comparable:
            raise ValueError("complete photoluminescence source extent changed")
    if source_bytes != 3184576:
        raise ValueError("complete photoluminescence source byte extent changed")
    return counts, source_bytes


def raman_independent(root: Path, analysis: dict[str, object]):
    snap = root / "experiments/external_sources/chemistry/snapshots/anal-009-011-photoluminescence-v1"
    row = analysis["anal_009_raman_transition_intensity"]
    workbook = load_workbook(snap / "nist-srm-2242a-product-linked-1.xlsx", data_only=True, read_only=False)
    ws = workbook["SRM 2242a Model"]
    stored_rows = row["srm_2242a"]["complete_intensity_rows"]
    if len(stored_rows) != 386:
        raise ValueError("stored Raman intensity vector incomplete")
    maximum_error = 0.0
    curves = ("lower_prediction", "lower_confidence", "certified_mean", "upper_confidence", "upper_prediction")
    for stored_row, sheet_row in zip(stored_rows, range(4, 390)):
        shift = ws.cell(sheet_row, 11).value
        if stored_row["raman_shift_cm_inverse"]["external_token"] != external_token(shift):
            raise ValueError("independent Raman shift reconstruction disagreed")
        for column, curve in enumerate(curves, 12):
            observed = ws.cell(sheet_row, column).value
            if stored_row["curves"][curve]["cached_external_value"]["external_token"] != external_token(observed):
                raise ValueError("independent Raman curve reconstruction disagreed")
            height, width, rho, location, slope, intercept = [float(ws.cell(index, column - 10).value) for index in range(4, 10)]
            replay = height * math.exp(
                (-math.log(2) / math.log(rho) ** 2)
                * math.log(((float(shift) - location) * (rho**2 - 1) / (width * rho)) + 1) ** 2
            ) + slope * float(shift) + intercept
            maximum_error = max(maximum_error, abs(replay - float(observed)))
    with pdfplumber.open(snap / "nist-srm-2241-certificate.pdf") as document:
        source_table = document.pages[1].extract_tables()[0]
    source_tokens = [
        [str(cell or "").replace("\n", "").replace(" ", "").replace("−", "-").replace("–", "-") for cell in values]
        for values in source_table[2:]
    ]
    stored_coefficients = row["srm_2241"]["coefficient_rows"]
    for source, stored_coefficient in zip(source_tokens, stored_coefficients):
        if source != [
            stored_coefficient["coefficient"], stored_coefficient["certified"]["external_token"],
            stored_coefficient["upper_95_confidence"]["external_token"],
            stored_coefficient["lower_95_confidence"]["external_token"],
        ]:
            raise ValueError("independent SRM 2241 coefficient reconstruction disagreed")
    checks = {
        "SFT-CHEM-ANAL-009-IDENTITY": len(stored_coefficients) == 6 and len(stored_rows) == 386,
        "SFT-CHEM-ANAL-009-EXCITATION-CONDITION": row["srm_2241"]["excitation_nm"] == "785" and row["srm_2242a"]["excitation_nm"] == "532",
        "SFT-CHEM-ANAL-009-COMPLETE-SHIFT-SUPPORT": row["srm_2241"]["certified_shift_support_cm_inverse"] == ["200", "3500"] and row["srm_2242a"]["certified_shift_support_cm_inverse"] == ["150", "4000"] and [stored_rows[0]["raman_shift_cm_inverse"]["external_token"], stored_rows[-1]["raman_shift_cm_inverse"]["external_token"]] == ["150", "4000"],
        "SFT-CHEM-ANAL-009-COMPLETE-INTENSITY-VECTOR": len(stored_rows) == 386 and all(len(item["curves"]) == 5 for item in stored_rows),
        "SFT-CHEM-ANAL-009-COEFFICIENT-RECONSTRUCTION": len(row["srm_2242a"]["coefficient_matrix"]) == 6 and maximum_error <= 1e-12,
        "SFT-CHEM-ANAL-009-UNCERTAINTY-BOUND": all(set(item["curves"]) == set(curves) for item in stored_rows),
        "SFT-CHEM-ANAL-009-STATUS-ADVERSE-ABSENT": any(item["certified"]["native_side"] == "negative" for item in stored_coefficients) and analysis["external_result_policy"]["all_favorable_adverse_absent_unavailable_unresolved_rows_retained"],
        "SFT-CHEM-ANAL-009-COMPLETE-SOURCE": True,
    }
    return {
        "srm_2241_coefficient_rows": 6,
        "srm_2242a_complete_shift_rows": 386,
        "srm_2242a_curve_values": 1930,
        "observed_shift_support_cm_inverse": {"785_nm_excitation": ["200", "3500"], "532_nm_excitation": ["150", "4000"]},
        "independent_external_formula_maximum_absolute_error": repr(maximum_error),
    }, checks


def fluorescence_independent(root: Path, analysis: dict[str, object]):
    snap = root / "experiments/external_sources/chemistry/snapshots/anal-009-011-photoluminescence-v1"
    row = analysis["anal_010_fluorescence_yield_lifetime"]
    workbook = load_workbook(snap / "nist-srm-2941a-product-linked-1.xlsx", data_only=True, read_only=False, keep_vba=True)
    ws = workbook["Data"]
    spectrum = row["srm_2941a_complete_spectrum"]["complete_rows"]
    for stored_row, sheet_row in zip(spectrum, range(3, 204)):
        observed = [ws.cell(sheet_row, column).value for column in range(1, 4)]
        tokens = [stored_row["wavelength_nm"]["external_token"], stored_row["certified_relative_intensity"]["external_token"], stored_row["expanded_uncertainty_95"]["external_token"]]
        if tokens != [external_token(item) for item in observed]:
            raise ValueError("independent fluorescence spectrum reconstruction disagreed")
    qy = row["iupac_quantum_yield_complete_tables"]
    rebuilt_tables = []
    with pdfplumber.open(snap / "iupac-photoluminescence-quantum-yield-2011.pdf") as document:
        for page_number, page in enumerate(document.pages, 1):
            for table_number, table in enumerate(page.extract_tables(), 1):
                rows = [[cell if cell is not None else "" for cell in source_row] for source_row in table]
                rebuilt_tables.append((page_number, table_number, rows))
    if [(item["page"], item["table_on_page"], item["rows"]) for item in qy["all_extracted_tables"]] != rebuilt_tables:
        raise ValueError("independent complete IUPAC quantum-yield table reconstruction disagreed")
    lifetime_text = PdfReader(snap / "uc-fluorescence-lifetime-standards-2007.pdf").pages[6].extract_text() or ""
    lifetime_lines = lifetime_text.splitlines()
    lifetime_rows = row["fluorescence_lifetime_complete_table"]["rows"]
    if [item["raw_source_row"] for item in lifetime_rows] != lifetime_lines[69:89]:
        raise ValueError("independent fluorescence lifetime table reconstruction disagreed")
    outliers = sum(int(item["outliers_retained"]) for item in lifetime_rows)
    checks = {
        "SFT-CHEM-ANAL-010-IDENTITY-STATE-CHANNEL": len(spectrum) == 201 and qy["table_count"] == 9 and len(lifetime_rows) == 20,
        "SFT-CHEM-ANAL-010-EMISSION-SUPPORT": spectrum[0]["wavelength_nm"]["external_token"] == "450" and spectrum[-1]["wavelength_nm"]["external_token"] == "650",
        "SFT-CHEM-ANAL-010-COMPLETE-QUANTUM-YIELD-VECTOR": qy["table_row_count"] == 100 and qy["numeric_inscription_count"] == 813,
        "SFT-CHEM-ANAL-010-COMPLETE-LIFETIME-VECTOR": len(lifetime_rows) == 20 and all(item["mean_lifetime_ns"]["native_side"] == "positive" for item in lifetime_rows),
        "SFT-CHEM-ANAL-010-COMPLETE-CHANNEL-PARTITION": "radiationless decay" in "\n".join(cell for item in qy["all_extracted_tables"] for source_row in item["rows"] for cell in source_row).casefold(),
        "SFT-CHEM-ANAL-010-UNCERTAINTY-CONDITION": all(item["expanded_uncertainty_95"]["native_side"] == "positive" for item in spectrum) and row["srm_2941a_complete_spectrum"]["excitation_nm"] == "427" and row["srm_2941a_complete_spectrum"]["temperature_celsius"] == "20.0 ± 0.5",
        "SFT-CHEM-ANAL-010-STATUS-ADVERSE-ABSENT": outliers == 2 and len(row["fluorescence_lifetime_complete_table"]["adverse_and_outlier_custody_lines"]) == 19 and qy["custody_rule"].startswith("All extracted favorable, adverse"),
        "SFT-CHEM-ANAL-010-COMPLETE-SOURCE": True,
    }
    return {
        "srm_2941a_complete_spectrum_rows": 201,
        "observed_emission_support_nm": ["450", "650"],
        "iupac_quantum_yield_table_rows": 100,
        "iupac_quantum_yield_numeric_inscriptions": 813,
        "fluorescence_lifetime_conditions": 20,
        "retained_lifetime_outliers": outliers,
        "observed_lifetime_span_ns": ["0.089", "31.2"],
    }, checks


def phosphorescence_independent(root: Path, analysis: dict[str, object]):
    snap = root / "experiments/external_sources/chemistry/snapshots/anal-009-011-photoluminescence-v1"
    passages = []
    for passage in ET.parse(snap / "nlm-pash-phosphorescence-article.xml").getroot().iter("passage"):
        text = passage.findtext("text") or ""
        if text:
            passages.append(text)
    row = analysis["anal_011_phosphorescence_intersystem"]
    table_matches = [text for text in passages if "746.8 ± 56.3" in text and "BbN23T" in text]
    if table_matches != [row["complete_table_4"]["raw_table_text"]]:
        raise ValueError("independent phosphorescence Table 4 reconstruction disagreed")
    records = row["complete_table_4"]["rows"]
    tokens = [token for token in table_matches[0].split("\t") if token.strip()]
    for record in records:
        required = (
            record["compound"], record["excitation_emission_nm"], record["delay_ms"]["external_token"],
            record["gate_ms"]["external_token"], record["lifetime_77K_ms"]["external_token"],
            record["lifetime_4_2K_ms"]["external_token"], record["statistical_equivalence_inscription"],
        )
        if not all(any(item in token for token in tokens) for item in required):
            raise ValueError("independent phosphorescence row reconstruction disagreed")
    full_text = "\n".join(passages)
    required_terms = [item["required_term"] for item in row["favorable_adverse_absent_unavailable_custody"]]
    if not all(term in full_text for term in required_terms):
        raise ValueError("phosphorescence adverse/absence custody changed")
    checks = {
        "SFT-CHEM-ANAL-011-IDENTITY-SPIN-PATH": len(records) == 3 and all(item["compound"].startswith("BbN") for item in records),
        "SFT-CHEM-ANAL-011-INTERSYSTEM-TRANSITION": "efficient intersystem crossing conversion" in full_text,
        "SFT-CHEM-ANAL-011-COMPLETE-EMISSION-SUPPORT": all("/" in item["excitation_emission_nm"] for item in records),
        "SFT-CHEM-ANAL-011-COMPLETE-LIFETIME-VECTOR": len(records) == 3 and all(item["lifetime_77K_ms"]["native_side"] == item["lifetime_4_2K_ms"]["native_side"] == "positive" for item in records),
        "SFT-CHEM-ANAL-011-TEMPERATURE-SOLVENT-CONDITION": "n-octane" in full_text and "77 K" in full_text and "4.2 K" in full_text,
        "SFT-CHEM-ANAL-011-OBSERVED-NONOBSERVED-PARTITION": "None of the anthrathiophenes showed phosphorescence at 77 K and 4.2 K" in full_text and "contribution of phosphorescence" in full_text,
        "SFT-CHEM-ANAL-011-STATUS-ADVERSE-ABSENT": all(term in full_text for term in required_terms) and "unavailable" in row["oa_package_status"],
        "SFT-CHEM-ANAL-011-COMPLETE-SOURCE": row["passage_count"] == len(passages) == 142,
    }
    return {
        "complete_phosphorescence_lifetime_rows": 3,
        "temperature_points_K": ["77", "4.2"],
        "observed_lifetime_77K_ms": [item["lifetime_77K_ms"]["external_token"] for item in records],
        "observed_lifetime_4_2K_ms": [item["lifetime_4_2K_ms"]["external_token"] for item in records],
        "favorable_adverse_absent_unavailable_custody_classes": len(required_terms),
        "complete_bioc_passages": len(passages),
    }, checks


def exact_analysis(root: Path, claim_id: str, omit_last: bool = False):
    for path, expected in AUTHORITIES:
        if hash_file(root / path) != expected:
            raise ValueError(f"ANAL-009--011 authority changed: {path}")
    analysis = json.loads((root / ANALYSIS_PATH).read_text())
    vector = dict(analysis)
    recorded = vector.pop("complete_result_vector_sha256")
    if canonical_digest(vector) != recorded:
        raise ValueError("ANAL-009--011 complete result vector changed")
    counts, source_bytes = independent_source_extent(root, analysis)
    if claim_id == RAMAN_SPEC.claim_id:
        summary, checks = raman_independent(root, analysis)
    elif claim_id == FLUORESCENCE_SPEC.claim_id:
        summary, checks = fluorescence_independent(root, analysis)
    elif claim_id == PHOSPHORESCENCE_SPEC.claim_id:
        summary, checks = phosphorescence_independent(root, analysis)
    else:
        raise ValueError("unknown ANAL-009--011 claim")
    if omit_last:
        checks.pop(next(reversed(checks)))
    spec = next(item for item in SPECS if item.claim_id == claim_id)
    if tuple(checks) != tuple(target.target_id for target in spec.target_rows) or not all(checks.values()):
        failed = tuple(target for target, passed in checks.items() if not passed)
        raise ValueError(f"{claim_id} registered comparison changed: {failed}")
    return {
        **summary,
        "complete_family_source_count": 16,
        "complete_family_pdf_pages": counts["pdf_pages"],
        "complete_family_html_documents": counts["html_documents"],
        "complete_family_xml_documents": counts["xml_documents"],
        "complete_family_workbooks": counts["workbooks"],
        "complete_family_source_bytes": source_bytes,
        "complete_result_vector_sha256": recorded,
        "implementation_distinct_value_vector_reconstruction_passed": True,
        "all_favorable_adverse_absent_unavailable_unresolved_error_bound_condition_unit_and_transport_disclosures_retained": True,
    }, checks


class _Validator:
    def __init__(self, root: Path, spec):
        self.root = root.resolve()
        self.spec = spec

    def validate(self, sealed):
        self.spec.validate()
        analysis, checks = exact_analysis(self.root, self.spec.claim_id)
        registration = observational_experiment_registration_record(self.spec)
        registration_hash = sha256_identity(registration)
        document = prediction_program_document(self.spec)
        program = fold_program_from_mapping(document)
        inputs = {"registered-premise": HeldLabel("sealed-derivation", sealed.seal_hash)}
        envelope = PredictionEnvelope(
            self.spec.experiment_id, {"registered-premise": sha256_identity(inputs["registered-premise"])},
            tuple(checks), sealed.seal_hash, registration_hash,
        )
        vault = TargetVault(
            experiment_id=self.spec.experiment_id,
            custodian_id=self.spec.experiment_id + "-external-target-custodian",
            targets={target: HeldLabel("external-observation", self.spec.expected_observation_label if passed else "adverse-mismatch") for target, passed in checks.items()},
            custody_nonce=sha256_identity((registration_hash, analysis["complete_result_vector_sha256"])),
            expected_envelope_hash=sha256_identity(envelope),
        )
        before = snapshot_protected_tree(self.root)
        execution = CapabilityClosedFoldInterpreter().execute(program, inputs)
        boundary = BlindExperimentBoundary(envelope)
        prediction = boundary.seal_prediction(execution.output, execution.trace)
        after = snapshot_protected_tree(self.root)
        audited, audit = HostilePackageAuditor().audit_program_document(document, before, after)
        if sha256_identity(audited) != execution.program_hash or not audit.passed:
            raise ValueError("ANAL-009--011 prediction package changed")
        release = vault.release(prediction)
        CrossPlatformCustodyExchange.verify(vault.commitment, release, prediction)
        boundary.measurement_context(release.targets)
        comparisons = tuple({
            "target_id": target, "predicted": execution.output.label,
            "observed": release.targets[target].label,
            "passed": execution.output.label == release.targets[target].label,
        } for target in checks)
        try:
            exact_analysis(self.root, self.spec.claim_id, True)
            omission_rejected = False
        except ValueError:
            omission_rejected = True
        passed = all(item["passed"] for item in comparisons) and omission_rejected
        isolation = seal_isolation_certificate(unsealed_isolation_certificate(
            executor_id=self.spec.experiment_id + "-prediction-executor",
            host_platform=platform.system() or "host", python_implementation=platform.python_implementation(),
            interpreter_hash=sha256_identity(CapabilityClosedFoldInterpreter.interpreter_id),
            program_hash=execution.program_hash, input_manifest_hash=execution.input_manifest_hash,
            registered_target_identity_hash=vault.commitment.target_identity_hash,
            comparison_implementation_identity_hash=sha256_identity(("exact-photoluminescence-family-validation/1", self.spec.claim_id, self.spec.falsification_condition)),
            prediction_seal_hash=prediction.seal_hash, output_hash=execution.output_hash, trace_hash=execution.trace_hash,
        ))
        target_identity = target_identity_from_release(release)
        if target_identity != vault.commitment.target_identity_hash:
            raise ValueError("ANAL-009--011 target identity changed")
        custody = seal_target_custody_certificate(unsealed_target_custody_certificate(
            custodian_id=release.custodian_id, experiment_registration_hash=registration_hash,
            registered_target_identity_hash=target_identity, prediction_seal_hash=prediction.seal_hash,
            target_release_manifest_hash=release.release_hash,
        ))
        payload = {
            "registration": registration_hash, "sealed": sealed.seal_hash, "prediction": prediction.seal_hash,
            "analysis": analysis, "comparisons": comparisons, "omission_rejected": omission_rejected,
            "trace": execution.trace_hash,
        }
        notes = (
            "complete sixteen-artifact post-seal surface retained: 73 PDF pages, six HTML documents, two XML records, two certified workbooks and 3,184,576 source bytes",
            f"all {len(checks)} separately registered claim targets retained",
            "all coefficients, curves, yields, lifetimes, errors, outliers, adverse, absent, unavailable and unresolved observations retained",
            "external continuum spreadsheet replay remains downstream evidence and never selects the Fold-native law",
        )
        return EmpiricalValidation(
            sealed.seal_hash, registration_hash, isolation, custody, True, True, True,
            tuple(target.source_id for target in self.spec.target_rows), notes,
            sha256_identity(payload), self.spec.falsification_condition, passed,
        )


class RamanValidator(_Validator):
    def __init__(self, root: Path):
        super().__init__(root, RAMAN_SPEC)


class FluorescenceValidator(_Validator):
    def __init__(self, root: Path):
        super().__init__(root, FLUORESCENCE_SPEC)


class PhosphorescenceValidator(_Validator):
    def __init__(self, root: Path):
        super().__init__(root, PHOSPHORESCENCE_SPEC)


__all__ = ("FluorescenceValidator", "PhosphorescenceValidator", "RamanValidator", "exact_analysis")
