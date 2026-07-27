#!/usr/bin/env python3
"""Normalize the complete post-seal KIN-010 catalytic-turnover source record."""

from __future__ import annotations

from io import BytesIO
import hashlib
import json
from pathlib import Path
import zipfile

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SNAPSHOT_ROOT = ROOT / "experiments/external_sources/chemistry/snapshots/kin-010-catalytic-turnover-v1"
SPEC_PATH = ROOT / "experiments/external_sources/chemistry/catalytic_turnover_capture_spec_v1.json"
SPEC_HASH = "sha256:e8874415767d9e257d94e860701dc839fdefd2761611a3a399b2885101e9a033"
INVENTORY_PATH = SNAPSHOT_ROOT / "source-inventory-v1.json"
INVENTORY_HASH = "sha256:1a70201cef55873701d19bae35487868c55fa23852cc251f98df0afec6ec9ee9"
IDENTITY_PATH = ROOT / "experiments/external_sources/chemistry/catalytic_turnover_target_identities_v1.json"
IDENTITY_HASH = "sha256:379360d1145dce4e4521525e60786e1ab39192f83a9b3cd9d95c13fdabdf7fb7"
TARGET_PATH = ROOT / "experiments/external_sources/chemistry/catalytic_turnover_withheld_targets_v1.json"
TARGET_HASH = "sha256:b6d7b668b74bb6d4ab3e367791554c788c10946ebe6d6020e3de8362c4a9b5f7"
OUTPUT_PATH = SNAPSHOT_ROOT / "catalytic-turnover-primary-records-v1.json"
FIGURE_6_ARCHIVE = "source-data-figure-6.zip"
SUPPLEMENTARY_DATA_ARCHIVE = "supplementary-data.zip"
TABLE_S1_MEMBER = "SI_source_data0708/TableS1/TableS1.png"
TABLE_S1_MEMBER_HASH = "sha256:70682f1043a7ff08682aa574a0b90c247065ff556a8eb1a10521a5f5f647daf8"


def sha_bytes(payload: bytes) -> str:
    return "sha256:" + hashlib.sha256(payload).hexdigest()


def sha_file(path: Path) -> str:
    return sha_bytes(path.read_bytes())


def external_table(rows: tuple[tuple[str, ...], ...], headers: tuple[str, ...]) -> dict:
    if any(len(row) != len(headers) for row in rows):
        raise ValueError("KIN-010 external table width changed")
    return {
        "headers": headers,
        "rows": tuple(dict(zip(headers, row)) for row in rows),
        "all_values_are_postseal_external_inscriptions_not_fold_proof_values": True,
    }


def main() -> None:
    for path, expected in (
        (SPEC_PATH, SPEC_HASH), (INVENTORY_PATH, INVENTORY_HASH),
        (IDENTITY_PATH, IDENTITY_HASH), (TARGET_PATH, TARGET_HASH),
    ):
        if sha_file(path) != expected:
            raise ValueError(f"KIN-010 sealed source boundary changed: {path}")

    inventory = json.loads(INVENTORY_PATH.read_text())
    identities = json.loads(IDENTITY_PATH.read_text())
    targets = json.loads(TARGET_PATH.read_text())
    identity_rows = tuple(identities["rows"])
    target_rows = tuple(targets["rows"])
    if (
        len(identity_rows) != len(target_rows) != 497
        or len(identity_rows) != 497
        or tuple(row["source_record_ordinal"] for row in identity_rows) != tuple(range(1, 498))
        or any(
            identity["target_id"] != target["target_id"]
            or identity["source_record_identity"] != target["source_record_identity"]
            for identity, target in zip(identity_rows, target_rows)
        )
    ):
        raise ValueError("KIN-010 complete source-ordered target vector changed")

    source_class_census: dict[str, int] = {}
    for row in target_rows:
        key = row["source_record_class"]
        source_class_census[key] = source_class_census.get(key, 0) + 1
    expected_source_class_census = {
        "complete-article-landing-record": 1,
        "attempted-article-pdf-returned-html-record": 1,
        "complete-supplementary-information-page": 106,
        "complete-supplementary-video": 1,
        "complete-zenodo-metadata-record": 1,
        "complete-source-data-archive-member": 387,
    }
    if source_class_census != expected_source_class_census:
        raise ValueError("KIN-010 complete source-class census changed")

    pages = {
        row["source_page_ordinal"]: " ".join(row["target_payload"]["complete_extracted_page_text"].split())
        for row in target_rows
        if row["source_record_class"] == "complete-supplementary-information-page"
    }
    required_page_fragments = {
        31: ("sequential electrical signals with the periodical pattern", "catalytic turnover rate became slower", "States 3 and 4 at 243 K"),
        33: ("LPd(0) (State 1)", "LPd(Ph)(OR) (State 3)", "pre -transmetalation intermediate ( State 4 )"),
        35: ("four conductivity states", "four intermediates in the catalytic cycle"),
        36: ("oxidative addition, and ligand exchange", "conductance state 4"),
        37: ("suppress the reductive elimination process", "conductance state 5", "four conductance states"),
        42: ("Substitute effects", "para-substituted PhB(OH) 2", "300 mV and 298 K"),
        43: ("Table S1", "corresponding catalytic cycle, TOF", "TOF value as the reaction rate"),
        44: ("number of catalytic cycles gradually decreased", "0%, 20%, 40%, 60%, 80%, and 100% toluene"),
        46: ("number of cycles per unit time decreased", "fewer catalytic cycles", "insufficient data"),
        47: ("maximum likelihood method", "transmetalation is the rate-determining step", "less reversible"),
        48: ("Table S2", "State 1 to State 4", "2.7×10−5", "8.1"),
        53: ("Table S3", "another set of data", "8.2×10−5", "1.0"),
        55: ("State 1 to 2", "State 2 to 3", "State 3 to 4", "State 4 to 5", "State 5 to 1"),
        60: ("under different bias voltages", "reverse reaction rates", "regulation of electrical field"),
    }
    for page, fragments in required_page_fragments.items():
        if page not in pages or any(fragment not in pages[page] for fragment in fragments):
            raise ValueError(f"KIN-010 decisive supplementary page changed: {page}")

    target_member = {
        (row.get("archive_identity"), row["source_record_identity"]): row["target_payload"]
        for row in target_rows if row["source_record_class"] == "complete-source-data-archive-member"
    }
    figure_6_path = SNAPSHOT_ROOT / FIGURE_6_ARCHIVE
    with zipfile.ZipFile(figure_6_path) as archive:
        members = {
            name: archive.read(name)
            for name in (
                "Fig. 6/fig.6a.txt", "Fig. 6/fig.6b.txt", "Fig. 6/fig.6c.txt",
                "Fig. 6/fig.6f.txt", "Fig. 6/Fig.6g.xlsx",
            )
        }
    for name, payload in members.items():
        registered = target_member[(FIGURE_6_ARCHIVE, name)]
        if sha_bytes(payload) != registered["complete_member_hash"] or len(payload) != registered["complete_member_byte_count"]:
            raise ValueError(f"KIN-010 Figure 6 member changed: {name}")

    raw_a = members["Fig. 6/fig.6a.txt"].decode().splitlines()
    raw_b = members["Fig. 6/fig.6b.txt"].decode().splitlines()
    if len(raw_a) != 287701 or len(raw_b) != 97916:
        raise ValueError("KIN-010 complete raw trace row census changed")
    figure_6c_lines = tuple(line.split() for line in members["Fig. 6/fig.6c.txt"].decode().splitlines() if line.strip())
    figure_6f_lines = tuple(line.split() for line in members["Fig. 6/fig.6f.txt"].decode().splitlines() if line.strip())
    if figure_6c_lines[0] != ["x", "y"] or len(figure_6c_lines) != 8:
        raise ValueError("KIN-010 Figure 6c exact external vector changed")
    if len(figure_6f_lines) != 8 or any(len(row) != 8 for row in figure_6f_lines):
        raise ValueError("KIN-010 Figure 6f complete histogram vector changed")

    workbook = load_workbook(BytesIO(members["Fig. 6/Fig.6g.xlsx"]), data_only=False, read_only=True)
    if workbook.sheetnames != ["Sheet1"]:
        raise ValueError("KIN-010 Figure 6g workbook topology changed")
    worksheet = workbook["Sheet1"]
    workbook_rows = tuple(tuple("EmptyOne" if cell is None else str(cell) for cell in row) for row in worksheet.iter_rows(values_only=True))
    if len(workbook_rows) != 7 or any(len(row) != 17 for row in workbook_rows):
        raise ValueError("KIN-010 Figure 6g complete workbook cell census changed")

    with zipfile.ZipFile(SNAPSHOT_ROOT / SUPPLEMENTARY_DATA_ARCHIVE) as archive:
        table_s1_image = archive.read(TABLE_S1_MEMBER)
    if sha_bytes(table_s1_image) != TABLE_S1_MEMBER_HASH:
        raise ValueError("KIN-010 visually inspected Table S1 image changed")
    if target_member[(SUPPLEMENTARY_DATA_ARCHIVE, TABLE_S1_MEMBER)]["complete_member_hash"] != TABLE_S1_MEMBER_HASH:
        raise ValueError("KIN-010 Table S1 image is not bound to the complete target vector")

    table_s1 = external_table(
        (
            ("OCH3", "-0.27", "0.5", "-1.9"),
            ("CH3", "-0.17", "4.6", "-0.9"),
            ("H", "0", "29.6", "0.0"),
            ("Cl", "0.23", "39.0", "3.1×10^-2"),
            ("COOMe", "0.45", "203.9", "0.7"),
            ("CF3", "0.54", "615.6", "1.2"),
            ("CN", "0.66", "2098.7", "1.8"),
        ),
        ("substituent", "sigma_p_external_inscription", "TOF_per_second_external_inscription", "lg_k_over_k0_external_inscription"),
    )
    table_s2 = external_table(
        (
            ("258 K", "2.7×10^-5", "1.5×10^-5"), ("278 K", "0.7", "8.1"),
            ("298 K", "5.9", "2.7×10^-4"), ("318 K", "1.8×10^-4", "1.3×10^-4"),
            ("338 K", "4.3×10^-3", "1.7×10^-3"),
        ),
        ("condition", "State_1_to_State_4_per_second_external_inscription", "State_4_to_State_1_per_second_external_inscription"),
    )
    table_s3 = external_table(
        (
            ("258 K", "8.2×10^-5", "1.2×10^-4"), ("278 K", "1.6×10^-3", "1.0"),
            ("298 K", "2.9", "1.8×10^-5"), ("318 K", "4.8×10^-3", "3.0×10^-4"),
            ("338 K", "2.3×10^-3", "7.9×10^-5"),
        ),
        ("condition", "State_1_to_State_4_per_second_external_inscription", "State_4_to_State_1_per_second_external_inscription"),
    )

    document = {
        "schema": "sft-v3-catalytic-turnover-primary-records/1",
        "claim_id": "SFT-CHEM-CATALYTIC-TURNOVER-CYCLE-FREQUENCY-010",
        "chemistry_obligation": "SFT-CHEM-OBL-KIN-010",
        "source_identity": {
            "article_doi": "10.1038/s41565-021-00959-4",
            "source_data_repository_doi": "10.5281/zenodo.4903414",
            "system": "single-palladium-molecule-Suzuki-Miyaura-complete-cycle-surface",
        },
        "sealed_boundaries": {
            "prefetch_specification": SPEC_HASH, "source_inventory": INVENTORY_HASH,
            "value_free_identity_registry": IDENTITY_HASH, "withheld_complete_target_registry": TARGET_HASH,
        },
        "complete_registered_target_count": 497,
        "complete_source_class_census": source_class_census,
        "complete_supplementary_page_count": 106,
        "complete_supplementary_movie_frame_count": 1604,
        "complete_archive_count": 7,
        "complete_archive_member_count": 387,
        "complete_source_file_count": inventory["complete_source_file_count"],
        "complete_source_file_byte_count": sum(row["byte_count"] for row in inventory["complete_source_files"]),
        "article_pdf_unavailable_html_response_retained_as_adverse_record": True,
        "structural_cycle": {
            "registered_state_count": 5,
            "separately_observed_conductance_state_count": 4,
            "ordered_states": (
                {"state": "State 1", "identity": "LPd(0) catalyst entry and return", "source_status": "separately-observed-conductance-state"},
                {"state": "State 2", "identity": "oxidative-addition product", "source_status": "structural-intermediate-not-separately-resolved-as-conductance-state"},
                {"state": "State 3", "identity": "LPd(Ph)(OR) ligand-exchange species", "source_status": "separately-observed-conductance-state"},
                {"state": "State 4", "identity": "pre-transmetalation intermediate", "source_status": "separately-observed-conductance-state"},
                {"state": "State 5", "identity": "species before reductive elimination", "source_status": "separately-observed-conductance-state"},
            ),
            "ordered_transition_word": (
                {"entry": "State 1", "exit": "State 2", "source_process": "oxidative addition"},
                {"entry": "State 2", "exit": "State 3", "source_process": "ligand exchange"},
                {"entry": "State 3", "exit": "State 4", "source_process": "pre-transmetalation"},
                {"entry": "State 4", "exit": "State 5", "source_process": "transmetalation"},
                {"entry": "State 5", "exit": "State 1", "source_process": "reductive elimination and exact catalyst return"},
            ),
            "entry_state_equals_return_state": True,
            "every_transition_occurrence_retained": True,
        },
        "table_s1_visually_inspected_image_member_hash": TABLE_S1_MEMBER_HASH,
        "complete_substituent_turnover_vector": table_s1,
        "independent_state_1_state_4_rate_vector_table_s2": table_s2,
        "independent_state_1_state_4_rate_vector_table_s3": table_s3,
        "table_s2_and_table_s3_remain_separate_without_selection_or_averaging": True,
        "figure_6_source_data": {
            "fig6a_complete_trace_row_count_including_header": len(raw_a),
            "fig6b_complete_trace_row_count_including_header": len(raw_b),
            "fig6a_member_hash": sha_bytes(members["Fig. 6/fig.6a.txt"]),
            "fig6b_member_hash": sha_bytes(members["Fig. 6/fig.6b.txt"]),
            "fig6c_complete_external_xy_vector": tuple({"x": row[0], "y": row[1]} for row in figure_6c_lines[1:]),
            "fig6f_complete_external_histogram_rows": figure_6f_lines,
            "fig6g_complete_external_workbook_rows": workbook_rows,
            "all_signed_decimal_and_zero_glyphs_are_external_inscriptions_only": True,
            "all_complete_raw_rows_remain_byte_bound_without_selection": True,
        },
        "complete_decisive_supplementary_page_ordinals": tuple(required_page_fragments),
        "low_temperature_fewer_cycle_and_insufficient_fit_data_adverse_record_retained": True,
        "source_reported_maximum_likelihood_single_exponent_Eyring_Arrhenius_Hess_and_other_fits_or_calculations_retained_as_postseal_provenance_only": True,
        "source_reported_rates_TOF_dwell_times_frequencies_conditions_errors_and_uncertainties_used_as_fold_proof_parameters": False,
        "imported_turnover_frequency_rate_equation_Michaelis_Menten_steady_state_stochastic_cycle_weight_fitted_efficiency_selection_average_interpolation_or_target_correction_used_in_law": False,
        "native_numerical_zero_negative_irrational_imaginary_signed_or_continuum_proof_value_used": False,
        "external_zero_and_negative_glyphs_preserved_only_as_source_inscriptions": True,
    }
    OUTPUT_PATH.write_text(json.dumps(document, indent=2, sort_keys=True, ensure_ascii=False) + "\n", encoding="utf-8")
    print(json.dumps({
        "output_path": str(OUTPUT_PATH.relative_to(ROOT)), "output_hash": sha_file(OUTPUT_PATH),
        "complete_target_count": 497, "complete_archive_member_count": 387,
        "complete_raw_trace_rows": len(raw_a) + len(raw_b), "turnover_value_rows": 7,
    }, indent=2))


if __name__ == "__main__":
    main()
