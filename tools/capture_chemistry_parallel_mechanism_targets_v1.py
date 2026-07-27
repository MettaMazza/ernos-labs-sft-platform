#!/usr/bin/env python3
"""Release the complete KIN-008 workbook only after the value-free identity seal."""

from __future__ import annotations

from decimal import Decimal
from fractions import Fraction
import hashlib
import json
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
SPEC_PATH = ROOT / "experiments/external_sources/chemistry/parallel_mechanism_capture_spec_v1.json"
SPEC_HASH = "sha256:f32b98d3cc4f02c02f01249b0f92ce799d1453ae04d1f8c9c107be6a509a6e89"
INVENTORY_PATH = ROOT / "experiments/external_sources/chemistry/snapshots/kin-008-parallel-mechanism-v1/source-inventory-v1.json"
INVENTORY_HASH = "sha256:a3c79878aeb0383a64d8bcf9242e9865c791c872ac50f59692348b978cead0d0"
IDENTITY_PATH = ROOT / "experiments/external_sources/chemistry/parallel_mechanism_target_identities_v1.json"
IDENTITY_HASH = "sha256:08d42e20f3e4fa66ff46f98d046e160e5a7375b32f7d6d036debddfe3f1b90ca"
WORKBOOK_PATH = ROOT / "experiments/external_sources/chemistry/snapshots/kin-008-parallel-mechanism-v1/source-data.xlsx"
WORKBOOK_HASH = "sha256:3f6dbcf377f4780aaec4f5a3c1431d4e758a775271d0a6d301bd74aca9087095"
TARGET_PATH = ROOT / "experiments/external_sources/chemistry/parallel_mechanism_withheld_targets_v1.json"


def sha_file(path: Path) -> str:
    return "sha256:" + hashlib.sha256(path.read_bytes()).hexdigest()


def exact_positive(inscription: str) -> str:
    value = Fraction(Decimal(inscription))
    if value <= 0:
        raise ValueError("KIN-008 exact-positive conversion received a non-positive source inscription")
    return str(value)


def observed_number(value: int | float) -> dict[str, object]:
    inscription = str(value)
    decimal = Decimal(inscription)
    if decimal == 0:
        return {
            "source_numeric_inscription": inscription,
            "source_value_class": "external-zero-glyph-observed-absence",
            "sft_interpretation": "structural-EmptyOne-observed-absence",
        }
    if decimal < 0:
        magnitude = str(-decimal)
        return {
            "source_numeric_inscription": inscription,
            "source_value_class": "external-signed-directional-inscription",
            "held_orientation": "below-source-reference",
            "exact_positive_magnitude_fraction": exact_positive(magnitude),
        }
    return {
        "source_numeric_inscription": inscription,
        "source_value_class": "exact-positive-observed-magnitude",
        "exact_positive_fraction": exact_positive(inscription),
    }


def cell_payload(formula_cell, cached_cell, coordinate: str) -> dict[str, object]:
    value = formula_cell.value
    base = {
        "cell_coordinate": coordinate,
        "source_data_type": getattr(formula_cell, "data_type", "n"),
        "source_number_format": getattr(formula_cell, "number_format", "General"),
    }
    if value is None:
        return {**base, "source_value_class": "structural-EmptyOne-source-cell-absence"}
    if getattr(formula_cell, "data_type", "n") == "f":
        cached = cached_cell.value
        result = observed_number(cached) if isinstance(cached, (int, float)) else {
            "source_value_class": "structural-EmptyOne-no-cached-formula-result" if cached is None else "source-cached-formula-label",
            "source_cached_inscription": None if cached is None else str(cached),
        }
        return {
            **base,
            "source_value_class": "source-reported-derived-formula-not-used-as-proof-parameter",
            "source_formula_inscription": str(value),
            "source_cached_result": result,
        }
    if isinstance(value, bool):
        return {**base, "source_value_class": "source-boolean-label", "source_label": str(value)}
    if isinstance(value, (int, float)):
        return {**base, **observed_number(value)}
    return {**base, "source_value_class": "held-source-label", "source_label": str(value)}


def main() -> None:
    for path, expected in (
        (SPEC_PATH, SPEC_HASH), (INVENTORY_PATH, INVENTORY_HASH),
        (IDENTITY_PATH, IDENTITY_HASH), (WORKBOOK_PATH, WORKBOOK_HASH),
    ):
        if sha_file(path) != expected:
            raise ValueError(f"KIN-008 sealed source changed: {path}")
    identities_document = json.loads(IDENTITY_PATH.read_text())
    identities = tuple(identities_document.get("rows", ()))
    if (
        identities_document.get("target_values_or_hashes_present") is not False
        or identities_document.get("complete_registered_target_count") != 28
        or len(identities) != 28
    ):
        raise ValueError("KIN-008 value-free identity seal changed")

    formulas = load_workbook(WORKBOOK_PATH, read_only=True, data_only=False)
    cached = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    rows = []
    totals = {"cells": 0, "EmptyOne": 0, "external_zero": 0, "positive": 0, "signed": 0, "formula": 0, "label": 0}
    for identity in identities:
        name = identity["source_sheet_identity"]
        formula_sheet = formulas[name]
        cached_sheet = cached[name]
        if (
            formula_sheet.max_row != identity["declared_max_row"]
            or formula_sheet.max_column != identity["declared_max_column"]
        ):
            raise ValueError(f"KIN-008 registered worksheet topology changed: {name}")
        cells = []
        for source_row in range(1, formula_sheet.max_row + 1):
            for source_column in range(1, formula_sheet.max_column + 1):
                payload = cell_payload(
                    formula_sheet.cell(source_row, source_column),
                    cached_sheet.cell(source_row, source_column),
                    f"{get_column_letter(source_column)}{source_row}",
                )
                cells.append(payload)
                totals["cells"] += 1
                value_class = payload["source_value_class"]
                if value_class.startswith("structural-EmptyOne"):
                    totals["EmptyOne"] += 1
                elif value_class == "external-zero-glyph-observed-absence":
                    totals["external_zero"] += 1
                elif value_class == "exact-positive-observed-magnitude":
                    totals["positive"] += 1
                elif value_class == "external-signed-directional-inscription":
                    totals["signed"] += 1
                elif value_class == "source-reported-derived-formula-not-used-as-proof-parameter":
                    totals["formula"] += 1
                else:
                    totals["label"] += 1
        rows.append({
            **identity,
            "complete_rectangular_cell_count": len(cells),
            "complete_rectangular_cells": cells,
        })
    if totals["cells"] != identities_document["complete_registered_rectangular_cell_position_count"]:
        raise ValueError("KIN-008 complete registered workbook cell census changed")
    document = {
        "schema": "sft-v3-parallel-mechanism-withheld-target-release/1",
        "claim_id": "SFT-CHEM-PARALLEL-MECHANISM-COMPOSITION-008",
        "chemistry_obligation": "SFT-CHEM-OBL-KIN-008",
        "prefetch_spec_hash": SPEC_HASH,
        "source_inventory_hash": INVENTORY_HASH,
        "identity_registry_hash": IDENTITY_HASH,
        "source_workbook_hash": WORKBOOK_HASH,
        "release_requires_complete_identity_and_prediction_seal": True,
        "complete_registered_target_count": len(rows),
        "complete_registered_rectangular_cell_position_count": totals["cells"],
        "complete_cell_class_census": totals,
        "all_twenty_eight_worksheets_and_every_registered_cell_position_preserved": True,
        "source_formulas_retained_but_never_used_as_fold_proof_parameters": True,
        "external_zero_glyphs_translate_only_to_structural_EmptyOne": True,
        "rows": rows,
    }
    TARGET_PATH.write_text(json.dumps(document, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    print(json.dumps({
        "target_path": str(TARGET_PATH.relative_to(ROOT)),
        "target_hash": sha_file(TARGET_PATH),
        "complete_registered_target_count": len(rows),
        "complete_cell_class_census": totals,
    }, indent=2))


if __name__ == "__main__":
    main()
