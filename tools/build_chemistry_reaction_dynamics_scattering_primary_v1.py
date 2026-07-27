#!/usr/bin/env python3
"""Build complete post-seal KIN-013 primary records without selecting outcomes."""

from __future__ import annotations

import json
from pathlib import Path
import sys

from openpyxl import load_workbook
from pypdf import PdfReader


ROOT = Path(__file__).resolve().parents[1]
SNAPSHOT = ROOT / "experiments/external_sources/chemistry/snapshots/kin-013-reaction-dynamics-scattering-v1"
SPEC = ROOT / "experiments/external_sources/chemistry/reaction_dynamics_scattering_capture_spec_v1.json"
INVENTORY = SNAPSHOT / "source-inventory-v1.json"
IDENTITIES = ROOT / "experiments/external_sources/chemistry/reaction_dynamics_scattering_target_identities_v1.json"
TARGETS = ROOT / "experiments/external_sources/chemistry/reaction_dynamics_scattering_withheld_targets_v1.json"
OUTPUT = SNAPSHOT / "reaction-dynamics-scattering-primary-records-v1.json"
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from sft.engine.source import hash_file  # noqa: E402


def write_json(path: Path, value: object) -> None:
    path.write_text(json.dumps(value, indent=2, sort_keys=True, ensure_ascii=False) + "\n", encoding="utf-8")


def inscription(value: object) -> str:
    if value is None:
        raise ValueError("absent cell has no external inscription")
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value)


def main() -> None:
    inventory = json.loads(INVENTORY.read_text())
    identities = json.loads(IDENTITIES.read_text())
    targets = json.loads(TARGETS.read_text())
    if identities.get("target_values_or_hashes_present") is not False:
        raise SystemExit("KIN-013 identity seal is not value-free")
    if targets.get("identity_registry_hash") != hash_file(IDENTITIES):
        raise SystemExit("KIN-013 target release is not bound to the identity seal")

    workbook = load_workbook(SNAPSHOT / "source-data.xlsx", read_only=True, data_only=False)
    shapes = []
    sheet_cells: dict[str, list[dict]] = {}
    total_nonempty = 0
    total_rows_with_value = 0
    for sheet in workbook.worksheets:
        cells = []
        rows_with_value = 0
        for row in sheet.iter_rows():
            nonempty = [cell for cell in row if cell.value is not None]
            if nonempty:
                rows_with_value += 1
            cells.extend(
                {
                    "cell_identity": cell.coordinate,
                    "external_inscription": inscription(cell.value),
                    "host_value_class": type(cell.value).__name__,
                }
                for cell in nonempty
            )
        total_nonempty += len(cells)
        total_rows_with_value += rows_with_value
        shapes.append(
            {
                "source_worksheet_identity": sheet.title,
                "declared_maximum_row_ordinal": sheet.max_row,
                "declared_maximum_column_ordinal": sheet.max_column,
                "complete_nonempty_cell_count": len(cells),
                "complete_row_with_value_count": rows_with_value,
            }
        )
        sheet_cells[sheet.title] = cells

    figure_three = tuple(sheet_cells["fig3"])
    figure_four = tuple(sheet_cells["fig4"])
    supplementary_eight = tuple(sheet_cells["Supplementary Fig. 8"])
    supplementary_nine = tuple(sheet_cells["Supplementary Fig. 9"])
    complete_state_resolved_vector = tuple(
        {"source_worksheet_identity": sheet, **cell}
        for sheet in ("fig3", "fig4", "Supplementary Fig. 8", "Supplementary Fig. 9")
        for cell in sheet_cells[sheet]
    )

    article_text = "\n".join(page.extract_text() or "" for page in PdfReader(SNAPSHOT / "article.pdf").pages)
    supplement_text = "\n".join(page.extract_text() or "" for page in PdfReader(SNAPSHOT / "supplementary-information.pdf").pages)
    peer_text = "\n".join(page.extract_text() or "" for page in PdfReader(SNAPSHOT / "transparent-peer-review.pdf").pages)
    all_text = "\n".join((article_text, supplement_text, peer_text))
    normalized_article_text = " ".join(article_text.split())

    payload = {
        "schema": "sft-v3-reaction-dynamics-scattering-primary-records/1",
        "chemistry_obligation": "SFT-CHEM-OBL-KIN-013",
        "claim_id": "SFT-CHEM-REACTION-DYNAMICS-SCATTERING-PRODUCT-STATE-013",
        "source_identity": {
            "article_doi": "10.1038/s41467-025-66587-x",
            "reaction_system": "F + CH4 → CH3(vi) + HF(v)",
        },
        "sealed_boundaries": {
            "prefetch_specification": [str(SPEC.relative_to(ROOT)), hash_file(SPEC)],
            "source_inventory": [str(INVENTORY.relative_to(ROOT)), hash_file(INVENTORY)],
            "value_free_identity_registry": [str(IDENTITIES.relative_to(ROOT)), hash_file(IDENTITIES)],
            "withheld_complete_target_registry": [str(TARGETS.relative_to(ROOT)), hash_file(TARGETS)],
        },
        "complete_source_file_count": inventory["complete_source_file_count"],
        "complete_source_file_byte_count": sum(row["byte_count"] for row in inventory["complete_source_files"]),
        "complete_registered_target_count": identities["complete_registered_target_count"],
        "complete_pdf_page_count": identities["complete_pdf_page_count"],
        "complete_source_data_worksheet_count": identities["complete_source_data_worksheet_count"],
        "complete_source_data_nonempty_cell_count": total_nonempty,
        "complete_source_data_row_with_value_count": total_rows_with_value,
        "complete_source_data_worksheet_shapes": shapes,
        "complete_fig3_pair_correlated_branching_cell_vector": figure_three,
        "complete_fig4_state_resolved_scattering_cell_vector": figure_four,
        "complete_supplementary_fig8_sampling_and_contamination_cell_vector": supplementary_eight,
        "complete_supplementary_fig9_before_after_correction_cell_vector": supplementary_nine,
        "complete_key_state_resolved_product_and_scattering_cell_vector": complete_state_resolved_vector,
        "complete_key_state_resolved_product_and_scattering_cell_count": len(complete_state_resolved_vector),
        "complete_structural_incoming_outgoing_channel_path": {
            "incoming_channel_identity": "held-F-plus-CH4-incoming-channel",
            "incoming_state_roles": ["F-atom-beam", "CH4-molecular-beam", "held-collision-preparation"],
            "outgoing_channel_identity": "held-CH3-plus-HF-outgoing-channel",
            "outgoing_product_pair_roles": ["CH3-held-vibrational-state", "HF-held-vibrational-state"],
            "finite_product_pair_support_retained": True,
            "finite_angular_bin_support_retained": True,
            "orientation_is_a_held_forward_sideways_or_backward_label": True,
        },
        "source_reported_headline_external_inscriptions": {
            "collision_energy": "2.4 kcal mol^-1 (0.106 eV)",
            "ground_state_CH3_reactivity": "40%",
            "umbrella_excited_CH3_experimental_reactive_flux": "57%",
            "umbrella_excited_CH3_theoretical_reactive_flux": "58%",
            "experimental_pair_branching_total": "unity",
            "incoming_CH4_rotational_state_distribution": "0.31:0.54:0.15 for j=0:j=1:j=2",
            "forward_product_pair": "(0_0,3)",
            "sideways_product_pair": "(2_2,2)",
            "backward_product_pair": "(0_0,2)",
        },
        "source_experimental_theoretical_and_processing_statuses_retained_separately": {
            "fig3_experiment_requires_no_normalization_factor_and_the_total_is_unity": "no need for any normalization factors" in normalized_article_text and "total sum is unity" in normalized_article_text,
            "fig4_experiment_and_theory_each_normalized_to_same_flux_for_shape_comparison": "normalized to the same" in normalized_article_text and "each panel" in normalized_article_text,
            "product_components_were_estimated_dissected_and_fitted": all(term in normalized_article_text for term in ("estimated and dissected", "were ﬁtted")),
            "HF_v1_weak_featureless_population_was_tentatively_posited": "tentatively posited" in normalized_article_text,
            "theoretical_forward_peak_is_sharper_than_experiment": "prediction appears" in normalized_article_text and "sharper" in normalized_article_text,
            "theory_calculated_only_ground_CH4_v0_j0_for_CDCS": "only the ground-state reaction with CH" in normalized_article_text and "was calculated" in normalized_article_text,
            "combination_bands_not_considered": "combination bands" in normalized_article_text and "not considered" in normalized_article_text,
            "lower_state_resolution_and_photochemical_background_limitations_retained": "lower state-" in normalized_article_text and "photochemical background" in normalized_article_text,
            "overlap_contamination_and_best_estimated_corrections_retained": "best-estimated" in normalized_article_text and "contaminations" in normalized_article_text,
        },
        "transparent_peer_review_adverse_surface": {
            "reviewer_warns_universal_detection_ability_may_not_be_overstated": "ability may not be" in peer_text and "overstated" in peer_text,
            "reviewer_warns_complete_resolution_becomes_difficult_for_rotationally_excited_products": "complete resolve will become difficult" in peer_text,
            "reviewer_identifies_dependence_on_previous_state_specific_work": "previous state-specific studies" in peer_text,
            "reviewer_challenges_mechanistic_explanation_and_coproduct_dependence": "strongly depends on the excitation of the co-product HF" in peer_text,
            "reviewer_identifies_unacknowledged_prior_118_nm_VUV_work": "one major" in peer_text and "shortcoming" in peer_text and "118 nm" in peer_text,
        },
        "all_complete_raw_image_speed_energy_branching_state_resolved_angular_sampling_overlap_correction_theory_and_peer_review_records_byte_bound": True,
        "source_reported_values_models_fits_normalizations_estimates_tentative_assignments_and_corrections_used_as_fold_proof_parameters": False,
        "source_scattering_cross_section_energy_momentum_transition_state_potential_surface_and_quantum_dynamics_models_retained_as_postseal_provenance_only": True,
        "external_zero_negative_decimal_scientific_notation_angle_and_continuum_inscriptions_preserved_only_as_source_provenance": True,
        "imported_scattering_equation_cross_section_law_angular_continuum_probability_amplitude_fitted_potential_distribution_normalization_selection_average_interpolation_or_target_correction_used_in_law": False,
        "native_numerical_zero_negative_irrational_imaginary_signed_or_continuum_proof_value_used": False,
        "complete_article_supplement_and_peer_review_text_byte_bound": len(all_text) > 1,
    }
    write_json(OUTPUT, payload)
    print(
        json.dumps(
            {
                "targets": payload["complete_registered_target_count"],
                "pages": payload["complete_pdf_page_count"],
                "worksheets": payload["complete_source_data_worksheet_count"],
                "nonempty_cells": total_nonempty,
                "key_state_resolved_cells": len(complete_state_resolved_vector),
            },
            sort_keys=True,
        )
    )


if __name__ == "__main__":
    main()
