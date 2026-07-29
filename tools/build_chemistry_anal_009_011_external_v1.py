#!/usr/bin/env python3
"""Independently reconstruct the complete ANAL-009--011 evidence vectors.

This reader is separate from the three Fold-native law implementations.  It
verifies every captured artifact, preserves every accessible page/document/
worksheet surface, and then reconstructs the registered Raman, fluorescence,
and phosphorescence targets without selecting favorable outcomes.

External signed and continuum inscriptions remain external evidence.  Their
native translations use a held side plus exact positive rational magnitude;
an external written zero denotes observed coincidence/absence (EmptyOne), not
a native numerical zero.
"""

from __future__ import annotations

from decimal import Decimal, InvalidOperation
from fractions import Fraction
import hashlib
import json
import math
from pathlib import Path
import re
import xml.etree.ElementTree as ET

from bs4 import BeautifulSoup
from openpyxl import load_workbook
import pdfplumber
from pypdf import PdfReader


ROOT = Path(__file__).resolve().parents[1]
SNAP = ROOT / "experiments/external_sources/chemistry/snapshots/anal-009-011-photoluminescence-v1"
INVENTORY = SNAP / "source-inventory-v1.json"
OUTPUT = SNAP / "complete-postseal-analysis-v1.json"

RAMAN_2241 = SNAP / "nist-srm-2241-certificate.pdf"
RAMAN_2242 = SNAP / "nist-srm-2242a-product-linked-1.xlsx"
FLUOR_2941 = SNAP / "nist-srm-2941a-product-linked-1.xlsx"
QY_PDF = SNAP / "iupac-photoluminescence-quantum-yield-2011.pdf"
LIFETIME_PDF = SNAP / "uc-fluorescence-lifetime-standards-2007.pdf"
PASH_XML = SNAP / "nlm-pash-phosphorescence-article.xml"


def digest(data: bytes) -> str:
    return "sha256:" + hashlib.sha256(data).hexdigest()


def canonical_digest(value: object) -> str:
    return digest(json.dumps(value, sort_keys=True, separators=(",", ":"), ensure_ascii=False).encode())


def cell_value(value: object) -> object:
    if value is None or isinstance(value, (str, int, bool)):
        return value
    if isinstance(value, float):
        return repr(value)
    return str(value)


def external_number(token: object, *, positive_side: str, negative_side: str) -> dict[str, object]:
    """Translate a written external decimal to a lawful exact Fold record."""
    written = str(token).strip().replace("−", "-").replace("–", "-").replace(" ", "")
    try:
        value = Decimal(written)
    except InvalidOperation as exc:
        raise SystemExit(f"non-decimal external inscription: {token!r}") from exc
    if not value.is_finite():
        raise SystemExit(f"non-finite external inscription: {token!r}")
    if value == 0:
        return {
            "external_token": str(token),
            "custody_status": "observed-coincidence-or-absence",
            "native_side": "coincident",
            "native_magnitude": "EmptyOne",
        }
    exact = Fraction(abs(value))
    return {
        "external_token": str(token),
        "custody_status": "measured",
        "native_side": positive_side if value > 0 else negative_side,
        "native_magnitude": f"{exact.numerator}/{exact.denominator}",
    }


def verify_inventory() -> dict[str, object]:
    inventory = json.loads(INVENTORY.read_text())
    payload = dict(inventory)
    stored_payload = payload.pop("inventory_payload_sha256")
    if canonical_digest(payload) != stored_payload:
        raise SystemExit("ANAL-009--011 inventory payload seal failed")
    if inventory.get("source_count") != 16 or len(inventory.get("sources", [])) != 16:
        raise SystemExit("ANAL-009--011 source surface is not the sealed sixteen-artifact set")
    for source in inventory["sources"]:
        path = ROOT / source["path"]
        data = path.read_bytes()
        if len(data) != source["byte_count"] or digest(data) != source["sha256"]:
            raise SystemExit(f"captured source changed: {source['path']}")
    return inventory


def workbook_surface(path: Path) -> dict[str, object]:
    formulas = load_workbook(path, data_only=False, read_only=True, keep_vba=True)
    cached = load_workbook(path, data_only=True, read_only=True, keep_vba=True)
    sheets: list[dict[str, object]] = []
    if len(formulas.worksheets) != len(cached.worksheets):
        raise SystemExit(f"workbook sheet-count mismatch: {path.name}")
    for formula_ws, cached_ws in zip(formulas.worksheets, cached.worksheets):
        if formula_ws.title != cached_ws.title:
            raise SystemExit(f"workbook sheet mismatch: {path.name}")
        cells = []
        formula_rows = list(formula_ws.iter_rows())
        cached_rows = list(cached_ws.iter_rows())
        if len(formula_rows) != len(cached_rows):
            raise SystemExit(f"workbook row-count mismatch: {path.name}:{formula_ws.title}")
        for formula_row, cached_row in zip(formula_rows, cached_rows):
            if len(formula_row) != len(cached_row):
                raise SystemExit(f"workbook column-count mismatch: {path.name}:{formula_ws.title}")
            for formula_cell, cached_cell in zip(formula_row, cached_row):
                if formula_cell.value is None and cached_cell.value is None:
                    continue
                cells.append(
                    {
                        "coordinate": formula_cell.coordinate,
                        "formula_or_inscription": cell_value(formula_cell.value),
                        "cached_value": cell_value(cached_cell.value),
                    }
                )
        sheets.append(
            {
                "title": formula_ws.title,
                "declared_max_row": formula_ws.max_row,
                "declared_max_column": formula_ws.max_column,
                "nonempty_cells": cells,
                "nonempty_cell_count": len(cells),
                "cell_vector_sha256": canonical_digest(cells),
            }
        )
    return {"path": path.relative_to(ROOT).as_posix(), "sheets": sheets}


def complete_source_surface(inventory: dict[str, object]) -> tuple[dict[str, object], dict[str, int]]:
    surface: dict[str, object] = {}
    counts = {"pdf_pages": 0, "html_documents": 0, "xml_documents": 0, "workbooks": 0, "characters": 0}
    for source in inventory["sources"]:
        path = ROOT / source["path"]
        data = path.read_bytes()
        record: dict[str, object] = {
            "path": source["path"],
            "media_kind": source["media_kind"],
            "byte_count": len(data),
            "sha256": digest(data),
        }
        kind = source["media_kind"]
        if kind == "pdf":
            page_vector = []
            for number, page in enumerate(PdfReader(path).pages, 1):
                text = page.extract_text() or ""
                page_vector.append(
                    {"page": number, "character_count": len(text), "text_sha256": digest(text.encode())}
                )
                counts["characters"] += len(text)
            record["complete_page_vector"] = page_vector
            counts["pdf_pages"] += len(page_vector)
        elif kind == "html":
            text = BeautifulSoup(data, "html.parser").get_text("\n")
            record["complete_document_text"] = {
                "character_count": len(text),
                "text_sha256": digest(text.encode()),
            }
            counts["html_documents"] += 1
            counts["characters"] += len(text)
        elif kind in {"bioc-xml", "pmc-oa-package-unavailable"}:
            root = ET.fromstring(data)
            texts = [element.text or "" for element in root.iter() if element.text]
            joined = "\n".join(texts)
            record["complete_xml_text"] = {
                "text_node_count": len(texts),
                "character_count": len(joined),
                "text_sha256": digest(joined.encode()),
            }
            if source.get("custody_status"):
                record["custody_status"] = source["custody_status"]
            counts["xml_documents"] += 1
            counts["characters"] += len(joined)
        elif kind == "linked-certified-workbook":
            record["complete_workbook_surface"] = workbook_surface(path)
            counts["workbooks"] += 1
        else:
            raise SystemExit(f"unsupported source kind: {kind}")
        surface[path.name] = record
    return surface, counts


def normalize_scientific_pdf_token(token: str) -> str:
    return token.replace("\n", "").replace(" ", "").replace("−", "-").replace("–", "-")


def raman_vector() -> dict[str, object]:
    with pdfplumber.open(RAMAN_2241) as document:
        tables = document.pages[1].extract_tables()
    if len(tables) != 1 or len(tables[0]) != 8:
        raise SystemExit("SRM 2241 coefficient table was not reconstructed exactly")
    coefficient_rows = []
    for row in tables[0][2:]:
        coefficient = normalize_scientific_pdf_token(row[0])
        values = [normalize_scientific_pdf_token(token) for token in row[1:4]]
        coefficient_rows.append(
            {
                "coefficient": coefficient,
                "certified": external_number(values[0], positive_side="positive", negative_side="negative"),
                "upper_95_confidence": external_number(values[1], positive_side="positive", negative_side="negative"),
                "lower_95_confidence": external_number(values[2], positive_side="positive", negative_side="negative"),
            }
        )

    # Normal mode is intentional here: repeated random access on a read-only
    # worksheet replays the XML stream for every cell and is not an independent
    # or efficient reconstruction of the finite matrix.
    formula_book = load_workbook(RAMAN_2242, data_only=False, read_only=False)
    value_book = load_workbook(RAMAN_2242, data_only=True, read_only=False)
    formula_ws = formula_book["SRM 2242a Model"]
    value_ws = value_book["SRM 2242a Model"]
    curves = ["lower_prediction", "lower_confidence", "certified_mean", "upper_confidence", "upper_prediction"]
    coefficients = ["H", "w", "rho", "x0", "m", "b"]
    coefficient_matrix = []
    for offset, name in enumerate(coefficients, 4):
        row = {"coefficient": name}
        for column, curve in enumerate(curves, 2):
            row[curve] = external_number(
                value_ws.cell(offset, column).value,
                positive_side="positive",
                negative_side="negative",
            )
        coefficient_matrix.append(row)

    intensity_rows = []
    maximum_absolute_reconstruction_error = 0.0
    for row_number in range(4, 390):
        shift = value_ws.cell(row_number, 11).value
        if shift is None:
            raise SystemExit(f"missing SRM 2242a Raman shift at row {row_number}")
        row: dict[str, object] = {
            "workbook_row": row_number,
            "raman_shift_cm_inverse": external_number(shift, positive_side="positive", negative_side="negative"),
            "curves": {},
        }
        for column, curve in enumerate(curves, 12):
            cached_value = value_ws.cell(row_number, column).value
            formula = formula_ws.cell(row_number, column).value
            if cached_value is None or not isinstance(formula, str) or not formula.startswith("="):
                raise SystemExit(f"incomplete SRM 2242a certified curve cell at row {row_number}, column {column}")
            params = [float(value_ws.cell(index, column - 10).value) for index in range(4, 10)]
            height, width, rho, location, slope, intercept = params
            independently_reconstructed = height * math.exp(
                (-math.log(2) / (math.log(rho) ** 2))
                * math.log(((float(shift) - location) * (rho**2 - 1) / (width * rho)) + 1) ** 2
            ) + slope * float(shift) + intercept
            error = abs(independently_reconstructed - float(cached_value))
            maximum_absolute_reconstruction_error = max(maximum_absolute_reconstruction_error, error)
            row["curves"][curve] = {
                "workbook_formula": formula,
                "cached_external_value": external_number(
                    cached_value, positive_side="positive", negative_side="negative"
                ),
                "independent_external_continuum_reconstruction": repr(independently_reconstructed),
                "absolute_reconstruction_error": repr(error),
            }
        intensity_rows.append(row)
    if len(intensity_rows) != 386 or maximum_absolute_reconstruction_error > 1e-12:
        raise SystemExit("SRM 2242a complete certified curve reconstruction failed")
    return {
        "srm_2241": {
            "excitation_nm": "785",
            "certified_shift_support_cm_inverse": ["200", "3500"],
            "temperature_celsius": ["20", "25"],
            "coefficient_rows": coefficient_rows,
        },
        "srm_2242a": {
            "excitation_nm": "532",
            "certified_shift_support_cm_inverse": ["150", "4000"],
            "temperature_celsius": ["20", "25"],
            "coefficient_matrix": coefficient_matrix,
            "complete_intensity_rows": intensity_rows,
            "intensity_row_count": len(intensity_rows),
            "independent_reconstruction_maximum_absolute_error": repr(maximum_absolute_reconstruction_error),
            "external_continuum_boundary": (
                "The NIST spreadsheet formula is independently replayed only as an external-measurement "
                "checker; it is not imported as a Fold-native arithmetic law."
            ),
        },
    }


LIFETIME_ROWS = [
    ("anthracene", "methanol", "5.1", "0.3", "6.1", "295-360", "375-442", "7", "7"),
    ("anthracene", "cyclohexane", "5.3", "0.1", "2.6", "295-360", "375-442", "7", "7"),
    ("9-cyanoanthracene", "methanol", "16", "1", "9.3", "295-360", "400-480", "7", "7"),
    ("9-cyanoanthracene", "cyclohexane", "12.7", "0.7", "5.5", "295-360", "400-450", "4", "4"),
    ("DPA", "methanol", "8.7", "0.5", "5.6", "295-360", "400-475", "8", "8"),
    ("DPA", "cyclohexane", "7.5", "0.4", "5.8", "295-360", "400-475", "8", "7"),
    ("N-methylcarbazole", "cyclohexane", "14.1", "0.9", "6.2", "290-325", "350-400", "6", "6"),
    ("coumarin 153", "methanol", "4.3", "0.2", "4.5", "295-442", "495-550", "5", "5"),
    ("erythrosin B", "water", "0.089", "0.003", "3.6", "488-568", "550-580", "6", "6"),
    ("erythrosin B", "methanol", "0.47", "0.02", "4.0", "488-568", "550-590", "6", "6"),
    ("NATA", "water", "3.1", "0.1", "3.6", "295-309", "330-410", "7", "7"),
    ("POPOP", "cyclohexane", "1.12", "0.04", "3.6", "295-360", "380-450", "8", "8"),
    ("PPO", "methanol", "1.65", "0.05", "2.7", "295-330", "340-400", "8", "8"),
    ("PPO", "cyclohexane", "1.36", "0.04", "2.6", "290-325", "360-450", "8", "8"),
    ("rhodamine B", "water", "1.74", "0.02", "0.9", "488-575", "560-630", "6", "5"),
    ("rhodamine B", "methanol", "2.5", "0.1", "4.0", "295,488-568", "550-630", "8", "8"),
    ("rubrene", "methanol", "9.9", "0.3", "3.2", "300,488,514", "550-610", "5", "5"),
    ("SPA", "water", "31.2", "0.4", "1.4", "300-330", "466-520", "5", "5"),
    ("p-terphenyl", "methanol", "1.17", "0.08", "6.5", "284-315", "330-380", "7", "7"),
    ("p-terphenyl", "cyclohexane", "0.98", "0.03", "3.3", "290-315", "330-390", "7", "7"),
]


def qy_tables() -> dict[str, object]:
    extracted = []
    with pdfplumber.open(QY_PDF) as document:
        for page_number, page in enumerate(document.pages, 1):
            for table_number, table in enumerate(page.extract_tables(), 1):
                rows = [[cell if cell is not None else "" for cell in row] for row in table]
                extracted.append(
                    {
                        "page": page_number,
                        "table_on_page": table_number,
                        "rows": rows,
                        "row_count": len(rows),
                        "row_vector_sha256": canonical_digest(rows),
                    }
                )
    if len(extracted) < 8 or sum(item["row_count"] for item in extracted) < 90:
        raise SystemExit("IUPAC quantum-yield table surface is incomplete")
    joined = "\n".join(cell for table in extracted for row in table["rows"] for cell in row)
    numeric_tokens = re.findall(r"(?<![A-Za-z])(?:\d+(?:\.\d+)?|\.\d+)(?:\s*±\s*\d+(?:\.\d+)?)?", joined)
    return {
        "all_extracted_tables": extracted,
        "table_count": len(extracted),
        "table_row_count": sum(item["row_count"] for item in extracted),
        "all_numeric_inscriptions": numeric_tokens,
        "numeric_inscription_count": len(numeric_tokens),
        "custody_rule": "All extracted favorable, adverse, contradictory, qualified, and unresolved table rows are retained.",
    }


def fluorescence_vector() -> dict[str, object]:
    formula_book = load_workbook(FLUOR_2941, data_only=False, read_only=False, keep_vba=True)
    value_book = load_workbook(FLUOR_2941, data_only=True, read_only=False, keep_vba=True)
    formula_ws = formula_book["Data"]
    value_ws = value_book["Data"]
    spectrum = []
    maximum_primary_duplicate_difference = 0.0
    for row_number in range(3, 204):
        wavelength, intensity, uncertainty = [value_ws.cell(row_number, column).value for column in range(1, 4)]
        duplicate = [value_ws.cell(row_number, column).value for column in range(8, 11)]
        if None in (wavelength, intensity, uncertainty) or None in duplicate or duplicate[0] != wavelength:
            raise SystemExit(f"SRM 2941a complete/duplicate certified vector is incomplete at row {row_number}")
        row_difference = max(abs(float(intensity) - float(duplicate[1])), abs(float(uncertainty) - float(duplicate[2])))
        maximum_primary_duplicate_difference = max(maximum_primary_duplicate_difference, row_difference)
        spectrum.append(
            {
                "workbook_row": row_number,
                "wavelength_nm": external_number(wavelength, positive_side="positive", negative_side="negative"),
                "certified_relative_intensity": external_number(
                    intensity, positive_side="positive", negative_side="negative"
                ),
                "expanded_uncertainty_95": external_number(
                    uncertainty, positive_side="positive", negative_side="negative"
                ),
                "primary_formula_or_inscription": [cell_value(formula_ws.cell(row_number, column).value) for column in range(1, 4)],
                "duplicate_formula_or_inscription": [cell_value(formula_ws.cell(row_number, column).value) for column in range(8, 11)],
                "primary_duplicate_maximum_absolute_difference": repr(row_difference),
            }
        )
    if len(spectrum) != 201 or spectrum[0]["wavelength_nm"]["external_token"] != "450" or spectrum[-1]["wavelength_nm"]["external_token"] != "650":
        raise SystemExit("SRM 2941a wavelength support is incomplete")

    lifetime_text = PdfReader(LIFETIME_PDF).pages[6].extract_text() or ""
    lines = lifetime_text.splitlines()
    raw_rows = lines[69:89]
    if len(raw_rows) != 20:
        raise SystemExit("fluorescence lifetime table did not yield twenty rows")
    structured_rows = []
    if len(raw_rows) != len(LIFETIME_ROWS):
        raise SystemExit("fluorescence lifetime source/specification row-count mismatch")
    for raw, values in zip(raw_rows, LIFETIME_ROWS):
        compound, solvent, mean, deviation, relative, excitation, emission, measured, used = values
        for required in (solvent, mean, deviation, relative, measured, used):
            if required not in raw:
                raise SystemExit(f"lifetime table transcription did not match source row: {raw}")
        structured_rows.append(
            {
                "raw_source_row": raw,
                "compound": compound,
                "solvent": solvent,
                "mean_lifetime_ns": external_number(mean, positive_side="positive", negative_side="negative"),
                "sample_standard_deviation_ns": external_number(
                    deviation, positive_side="positive", negative_side="negative"
                ),
                "relative_standard_deviation_percent": external_number(
                    relative, positive_side="positive", negative_side="negative"
                ),
                "excitation_nm_inscription": excitation,
                "emission_nm_inscription": emission,
                "measurements_recorded": measured,
                "measurements_used": used,
                "outliers_retained": str(int(measured) - int(used)),
            }
        )
    adverse_lines = lines[45:59] + lines[89:94]
    return {
        "srm_2941a_complete_spectrum": {
            "excitation_nm": "427",
            "temperature_celsius": "20.0 ± 0.5",
            "bandwidth_nm": "3.0",
            "complete_rows": spectrum,
            "row_count": len(spectrum),
            "primary_duplicate_wavelength_identity": True,
            "primary_duplicate_maximum_absolute_difference": repr(maximum_primary_duplicate_difference),
            "primary_duplicate_agree_within_external_binary_workbook_tolerance_1e_12": (
                maximum_primary_duplicate_difference <= 1e-12
            ),
        },
        "iupac_quantum_yield_complete_tables": qy_tables(),
        "fluorescence_lifetime_complete_table": {
            "temperature_celsius": "20",
            "rows": structured_rows,
            "row_count": len(structured_rows),
            "adverse_and_outlier_custody_lines": adverse_lines,
            "adverse_and_outlier_custody_sha256": canonical_digest(adverse_lines),
        },
    }


def bioc_passages() -> list[str]:
    passages = []
    for passage in ET.parse(PASH_XML).getroot().iter("passage"):
        text = passage.findtext("text") or ""
        if text:
            passages.append(text)
    return passages


def phosphorescence_vector() -> dict[str, object]:
    passages = bioc_passages()
    table_texts = [text for text in passages if "746.8 ± 56.3" in text and "BbN23T" in text]
    if len(table_texts) != 1:
        raise SystemExit("phosphorescence lifetime Table 4 was not uniquely reconstructed")
    table = table_texts[0]
    row_pattern = re.compile(
        r"(BbN(?:12|21|23)T)\t([^\t]+)\t([^\t]+)\t([^\t]+)\t([^\t]+)\t([^\t]+)\t([^\t]+)"
    )
    parsed = row_pattern.findall(table)
    if len(parsed) != 3:
        raise SystemExit("phosphorescence lifetime Table 4 did not yield three complete rows")
    rows = []
    for compound, wavelength, delay, gate, at_77, at_4_2, equivalence in parsed:
        mean_77, sd_77 = [part.strip() for part in at_77.split("±")]
        mean_4_2, sd_4_2 = [part.strip() for part in at_4_2.split("±")]
        rows.append(
            {
                "compound": compound,
                "excitation_emission_nm": wavelength,
                "delay_ms": external_number(delay, positive_side="positive", negative_side="negative"),
                "gate_ms": external_number(gate, positive_side="positive", negative_side="negative"),
                "lifetime_77K_ms": external_number(mean_77, positive_side="positive", negative_side="negative"),
                "lifetime_77K_standard_deviation_ms": external_number(
                    sd_77, positive_side="positive", negative_side="negative"
                ),
                "lifetime_4_2K_ms": external_number(mean_4_2, positive_side="positive", negative_side="negative"),
                "lifetime_4_2K_standard_deviation_ms": external_number(
                    sd_4_2, positive_side="positive", negative_side="negative"
                ),
                "statistical_equivalence_inscription": equivalence,
            }
        )
    custody_terms = (
        "No phosphorescence was observed from any of the studied compounds at RT",
        "None of the anthrathiophenes showed phosphorescence at 77 K and 4.2 K",
        "efficient intersystem crossing conversion",
        "stronger fluorescence than phosphorescence",
        "All phosphorescence decays showed well-behaved single exponential decays",
        "None of the studied compounds showed phosphorescence at RT, 77 K or 4.2 K",
        "lack of commercial standards prevented us",
    )
    custody = []
    for term in custody_terms:
        matches = [
            {"passage_index": index, "text": text}
            for index, text in enumerate(passages)
            if term in text
        ]
        if not matches:
            raise SystemExit(f"missing phosphorescence favorable/adverse custody term: {term}")
        custody.append({"required_term": term, "matching_passages": matches})
    return {
        "complete_table_4": {
            "raw_table_text": table,
            "rows": rows,
            "row_count": len(rows),
            "measurement_count_per_temperature": "3",
            "excitation_and_emission_bandpass_nm": "4",
        },
        "favorable_adverse_absent_unavailable_custody": custody,
        "complete_bioc_passage_vector": [
            {"passage_index": index, "character_count": len(text), "text_sha256": digest(text.encode())}
            for index, text in enumerate(passages)
        ],
        "passage_count": len(passages),
        "oa_package_status": "unavailable as declared by the captured PMC manifest; full BioC text retained",
    }


def main() -> None:
    if OUTPUT.exists():
        raise SystemExit("ANAL-009--011 complete analysis exists; rebuild prohibited")
    inventory = verify_inventory()
    surface, surface_counts = complete_source_surface(inventory)
    result = {
        "schema": "sft-v3-complete-postseal-analysis/1",
        "family": "ANAL-009-011-RAMAN-FLUORESCENCE-PHOSPHORESCENCE",
        "source_inventory_path": INVENTORY.relative_to(ROOT).as_posix(),
        "source_inventory_file_sha256": digest(INVENTORY.read_bytes()),
        "source_inventory_payload_sha256": inventory["inventory_payload_sha256"],
        "complete_source_surface": surface,
        "complete_source_surface_counts": surface_counts,
        "anal_009_raman_transition_intensity": raman_vector(),
        "anal_010_fluorescence_yield_lifetime": fluorescence_vector(),
        "anal_011_phosphorescence_intersystem": phosphorescence_vector(),
        "external_result_policy": {
            "source_selection_after_outcome_access": False,
            "all_favorable_adverse_absent_unavailable_unresolved_rows_retained": True,
            "external_signed_and_continuum_values_are_not_native_arithmetic": True,
            "external_zero_translates_to_EmptyOne_not_numerical_zero": True,
            "candidate_or_survivor_selection_from_measurements": False,
        },
    }
    result["complete_result_vector_sha256"] = canonical_digest(result)
    OUTPUT.write_text(json.dumps(result, indent=2, sort_keys=True, ensure_ascii=False) + "\n")
    print(json.dumps({
        "output": OUTPUT.relative_to(ROOT).as_posix(),
        "file_sha256": digest(OUTPUT.read_bytes()),
        "complete_result_vector_sha256": result["complete_result_vector_sha256"],
        "surface_counts": surface_counts,
        "raman_rows": 386,
        "fluorescence_spectrum_rows": 201,
        "fluorescence_lifetime_rows": 20,
        "phosphorescence_lifetime_rows": 3,
    }, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()
