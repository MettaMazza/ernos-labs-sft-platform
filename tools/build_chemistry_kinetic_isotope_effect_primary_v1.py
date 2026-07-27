#!/usr/bin/env python3
"""Normalize the complete post-seal KIN-012 source surface without fitting."""

from __future__ import annotations

import json
from pathlib import Path
import sys

from openpyxl import load_workbook
from pypdf import PdfReader


ROOT = Path(__file__).resolve().parents[1]
SNAPSHOT = ROOT / "experiments/external_sources/chemistry/snapshots/kin-012-kinetic-isotope-effect-v1"
SPEC = ROOT / "experiments/external_sources/chemistry/kinetic_isotope_effect_capture_spec_v1.json"
INVENTORY = SNAPSHOT / "source-inventory-v1.json"
IDENTITIES = ROOT / "experiments/external_sources/chemistry/kinetic_isotope_effect_target_identities_v1.json"
TARGETS = ROOT / "experiments/external_sources/chemistry/kinetic_isotope_effect_withheld_targets_v1.json"
OUTPUT = SNAPSHOT / "kinetic-isotope-effect-primary-records-v1.json"
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from sft.engine.source import hash_file  # noqa: E402


def inscription(value: object) -> str:
    if isinstance(value, bool) or value is None:
        raise ValueError("KIN-012 external rate inscription is absent or nonnumeric")
    return str(value)


def write_json(path: Path, value: object) -> None:
    path.write_text(json.dumps(value, indent=2, sort_keys=True, ensure_ascii=False) + "\n", encoding="utf-8")


def triplicate_rows(sheet, system: str, start: int = 2, temperature_column: int = 2, first_value_column: int = 3) -> list[dict]:
    rows = []
    for source_row in range(start, start + 5):
        temperature = inscription(sheet.cell(source_row, temperature_column).value)
        for replicate in range(1, 4):
            rows.append({
                "system_identity": system,
                "temperature_C_external_inscription": temperature,
                "replicate_ordinal": replicate,
                "rate_ratio_external_inscription": inscription(sheet.cell(source_row, first_value_column + replicate - 1).value),
                "source_worksheet_identity": sheet.title,
                "source_row_ordinal": source_row,
            })
    return rows


def single_rows(sheet, system: str, start: int, temperature_column: int, value_column: int) -> list[dict]:
    return [
        {
            "system_identity": system,
            "temperature_C_external_inscription": inscription(sheet.cell(source_row, temperature_column).value),
            "replicate_ordinal": 1,
            "rate_ratio_external_inscription": inscription(sheet.cell(source_row, value_column).value),
            "source_worksheet_identity": sheet.title,
            "source_row_ordinal": source_row,
        }
        for source_row in range(start, start + 5)
    ]


def main() -> None:
    inventory = json.loads(INVENTORY.read_text())
    identities = json.loads(IDENTITIES.read_text())
    targets = json.loads(TARGETS.read_text())
    if len(identities["rows"]) != 71 or len(targets["rows"]) != 71:
        raise ValueError("KIN-012 complete target boundary changed")
    workbook = load_workbook(SNAPSHOT / "source-data.xlsx", read_only=False, data_only=True)
    vector = []
    vector.extend(triplicate_rows(workbook["Figure 3a"], "Pt-TiO2-water-splitting-H2-H2O-over-D2O"))
    vector.extend(triplicate_rows(workbook["Figure 3b"], "Pt-TiO2-CO2-reduction-CO-H2O-over-D2O"))
    vector.extend(triplicate_rows(workbook["Figure 3c"], "anatase-TiO2-CO2-reduction-CO-H2O-over-D2O"))
    vector.extend(single_rows(workbook["Supplementary Figure 3"], "predeposited-Pt-TiO2-water-splitting-H2-H2O-over-D2O", 2, 2, 3))
    vector.extend(triplicate_rows(workbook["Supplementary Figure 4"], "Pt-TiO2-CO2-reduction-H2-H2O-over-D2O"))
    vector.extend(single_rows(workbook["Supplementary Figure 5"], "anatase-TiO2-vapour-CO2-reduction-CO-H2O-over-D2O", 2, 2, 3))
    vector.extend(single_rows(workbook["Supplementary Figure 6"], "anatase-TiO2-crystal-control-CO-H2O-over-D2O", 3, 8, 9))
    vector.extend(single_rows(workbook["Supplementary Figure 6"], "rutile-TiO2-crystal-control-CO-H2O-over-D2O", 3, 13, 14))
    vector.extend(single_rows(workbook["Supplementary Figure 8"], "001-facet-TiO2-control-CO-H2O-over-D2O", 3, 8, 9))
    vector.extend(single_rows(workbook["Supplementary Figure 9"], "oxygen-deficient-TiO2-control-CO-H2O-over-D2O", 3, 14, 15))
    shapes = []
    complete_nonempty_cells = 0
    complete_rows_with_values = 0
    for sheet in workbook.worksheets:
        nonempty = 0
        rows_with_values = 0
        for row in sheet.iter_rows(values_only=True):
            present = sum(value is not None for value in row)
            nonempty += present
            rows_with_values += present > 0
        complete_nonempty_cells += nonempty
        complete_rows_with_values += rows_with_values
        shapes.append({
            "source_worksheet_identity": sheet.title,
            "declared_maximum_row_ordinal": sheet.max_row,
            "declared_maximum_column_ordinal": sheet.max_column,
            "complete_nonempty_cell_count": nonempty,
            "complete_row_with_value_count": rows_with_values,
        })
    pdf_text = {
        name: "\n".join(page.extract_text() or "" for page in PdfReader(SNAPSHOT / name).pages)
        for name in ("article.pdf", "supplementary-information.pdf", "reporting-summary.pdf")
    }
    article = pdf_text["article.pdf"]
    reporting = pdf_text["reporting-summary.pdf"]
    source_files = tuple(inventory["complete_source_files"])
    payload = {
        "schema": "sft-v3-kinetic-isotope-effect-primary-records/1",
        "chemistry_obligation": "SFT-CHEM-OBL-KIN-012",
        "claim_id": "SFT-CHEM-KINETIC-ISOTOPE-EFFECT-RELATION-012",
        "source_identity": {
            "article_doi": "10.1038/s41467-024-44753-x",
            "system": "TiO2-H2O-D2O-CO2-complete-kinetic-isotope-effect-surface",
        },
        "sealed_boundaries": {
            "prefetch_specification": [str(SPEC.relative_to(ROOT)), hash_file(SPEC)],
            "source_inventory": [str(INVENTORY.relative_to(ROOT)), hash_file(INVENTORY)],
            "value_free_identity_registry": [str(IDENTITIES.relative_to(ROOT)), hash_file(IDENTITIES)],
            "withheld_complete_target_registry": [str(TARGETS.relative_to(ROOT)), hash_file(TARGETS)],
        },
        "complete_source_file_count": len(source_files),
        "complete_source_file_byte_count": sum(row["byte_count"] for row in source_files),
        "complete_registered_target_count": len(identities["rows"]),
        "complete_pdf_page_count": identities["complete_pdf_page_count"],
        "complete_source_data_worksheet_count": len(shapes),
        "complete_source_data_worksheet_shapes": shapes,
        "complete_source_data_nonempty_cell_count": complete_nonempty_cells,
        "complete_source_data_row_with_value_count": complete_rows_with_values,
        "complete_explicit_rate_ratio_vector_count": len(vector),
        "complete_explicit_rate_ratio_vector": vector,
        "source_reported_direct_decay_KIE_external_inscriptions": [
            {"path_identity": "O-H-over-O-D-water-splitting-decay", "KIE_external_inscription": "2.11"},
            {"path_identity": "O-H-over-O-D-CO2-present-decay", "KIE_external_inscription": "0.827"},
            {"path_identity": "O=C=O-H-plus-over-O=C=O-D-plus-decay", "KIE_external_inscription": "0.55"}
        ],
        "source_reported_temperature_series_boundary_external_inscriptions": {
            "temperatures_C": ["3", "6", "9", "12", "15"],
            "inverse_CO_KSIE_range": ["0.2", "0.9"],
            "water_splitting_KSIE_at_15_C": "2.8",
        },
        "complete_structural_isotopologue_path": {
            "light_isotopologue_identity": "H2O-held-source-label",
            "heavy_isotopologue_identity": "D2O-held-source-label",
            "same_reaction_path_roles_retained": True,
            "ordered_path_roles": ["reactant-and-condition-entry", "protonated-intermediate", "rate-determining-decay-or-production-event", "product-observation"],
            "every_rate_ratio_keeps_system_temperature_replicate_path_and_isotopologue_order": True,
        },
        "three_independent_experiments_retained_without_averaging": "Error bar represents three independent experiments" in article and all(row["replicate_ordinal"] in (1, 2, 3) for row in vector),
        "reporting_summary_admits_in_situ_infrared_is_not_standalone_evidence": "in-situ infrared data are not standalone evidence" in reporting,
        "reporting_summary_retains_reviewer_challenges_and_control_requests": all(phrase in reporting for phrase in ("Control-DRIFTS-experiments", "Could the authors please explain", "Very different KIEs are observed")),
        "source_interpretive_transition_state_zero_point_Hooke_quantum_calculation_and_fit_models_retained_as_postseal_provenance_only": all(phrase in article for phrase in ("Hooke", "Quantum chemical calculations", "transition state")),
        "source_reported_rate_ratio_decay_production_temperature_uncertainty_replicate_and_condition_values_used_as_fold_proof_parameters": False,
        "imported_KIE_equation_mass_frequency_law_transition_state_continuum_fitted_exponent_statistical_weight_selection_average_interpolation_or_target_correction_used_in_law": False,
        "native_numerical_zero_negative_irrational_imaginary_signed_or_continuum_proof_value_used": False,
        "external_zero_negative_decimal_and_continuum_inscriptions_preserved_only_as_source_provenance": True,
    }
    if (
        payload["complete_registered_target_count"] != 71
        or payload["complete_pdf_page_count"] != 47
        or payload["complete_source_data_worksheet_count"] != 23
        or payload["complete_explicit_rate_ratio_vector_count"] != 90
        or payload["three_independent_experiments_retained_without_averaging"] is not True
        or payload["reporting_summary_admits_in_situ_infrared_is_not_standalone_evidence"] is not True
        or payload["reporting_summary_retains_reviewer_challenges_and_control_requests"] is not True
    ):
        raise ValueError("KIN-012 complete primary normalization failed")
    write_json(OUTPUT, payload)
    print(json.dumps({
        "targets": payload["complete_registered_target_count"],
        "pages": payload["complete_pdf_page_count"],
        "worksheets": payload["complete_source_data_worksheet_count"],
        "rate_ratio_rows": payload["complete_explicit_rate_ratio_vector_count"],
        "nonempty_cells": payload["complete_source_data_nonempty_cell_count"],
    }, sort_keys=True))


if __name__ == "__main__":
    main()
