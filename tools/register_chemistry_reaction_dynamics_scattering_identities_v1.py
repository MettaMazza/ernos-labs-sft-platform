#!/usr/bin/env python3
"""Seal value-free KIN-013 source identities before target extraction."""

from __future__ import annotations

import json
from pathlib import Path
import sys

from openpyxl import load_workbook
from pypdf import PdfReader


ROOT = Path(__file__).resolve().parents[1]
SNAPSHOT = ROOT / "experiments/external_sources/chemistry/snapshots/kin-013-reaction-dynamics-scattering-v1"
SPEC = ROOT / "experiments/external_sources/chemistry/reaction_dynamics_scattering_capture_spec_v1.json"
INVENTORY = SNAPSHOT / "source-inventory-v1.json"
OUTPUT = ROOT / "experiments/external_sources/chemistry/reaction_dynamics_scattering_target_identities_v1.json"
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from sft.engine.source import hash_file  # noqa: E402


SOURCE_ID = "NATURE-COMMUNICATIONS-S41467-025-66587-X-COMPLETE"
SYSTEM = "F-CH4-to-CH3-HF-complete-pair-correlated-product-state-scattering-surface"


def write_json(path: Path, value: object) -> None:
    path.write_text(json.dumps(value, indent=2, sort_keys=True, ensure_ascii=False) + "\n", encoding="utf-8")


def common(document: str, record_class: str, record_identity: str) -> dict:
    return {
        "source_id": SOURCE_ID,
        "article_doi": "10.1038/s41467-025-66587-x",
        "incoming_outgoing_reaction_system_identity": SYSTEM,
        "source_document_identity": document,
        "source_record_class": record_class,
        "source_record_identity": record_identity,
    }


def main() -> None:
    rows = [common("article.html", "complete-article-landing-record", "complete-article-landing-html")]
    pdf_records = (
        ("article.pdf", "complete-article-pdf-page"),
        ("supplementary-information.pdf", "complete-supplementary-information-page"),
        ("transparent-peer-review.pdf", "complete-transparent-peer-review-page"),
    )
    page_counts = {}
    for document, record_class in pdf_records:
        count = len(PdfReader(SNAPSHOT / document).pages)
        page_counts[document] = count
        rows.extend(
            common(document, record_class, f"{document}-page-{page}") | {"source_page_ordinal": page}
            for page in range(1, count + 1)
        )
    workbook = load_workbook(SNAPSHOT / "source-data.xlsx", read_only=True, data_only=False)
    sheet_records = []
    for ordinal, sheet in enumerate(workbook.worksheets, start=1):
        sheet_records.append(
            common("source-data.xlsx", "complete-source-data-worksheet", sheet.title)
            | {
                "source_worksheet_ordinal": ordinal,
                "source_worksheet_identity": sheet.title,
                "source_workbook_member_identity": sheet._worksheet_path,
                "declared_maximum_row_ordinal": sheet.max_row,
                "declared_maximum_column_ordinal": sheet.max_column,
            }
        )
    rows.extend(sheet_records)
    rows = tuple(
        row
        | {
            "target_id": f"SFT-CHEM-KIN013-COMPLETE-SOURCE-RECORD-{ordinal:03d}",
            "source_record_ordinal": ordinal,
        }
        for ordinal, row in enumerate(rows, start=1)
    )
    payload = {
        "schema": "sft-v3-reaction-dynamics-scattering-target-identities/1",
        "chemistry_obligation": "SFT-CHEM-OBL-KIN-013",
        "claim_id": "SFT-CHEM-REACTION-DYNAMICS-SCATTERING-PRODUCT-STATE-013",
        "prefetch_spec_hash": hash_file(SPEC),
        "source_inventory_hash": hash_file(INVENTORY),
        "complete_registered_target_count": len(rows),
        "complete_pdf_page_count": sum(page_counts.values()),
        "complete_article_pdf_page_count": page_counts["article.pdf"],
        "complete_supplementary_information_page_count": page_counts["supplementary-information.pdf"],
        "complete_transparent_peer_review_page_count": page_counts["transparent-peer-review.pdf"],
        "complete_source_data_worksheet_count": len(sheet_records),
        "target_values_or_hashes_present": False,
        "all_incoming_outgoing_channel_product_state_angle_speed_energy_branching_experimental_theoretical_fit_normalization_estimate_tentative_background_control_limitation_adverse_reviewer_status_value_and_target_hash_values_absent": True,
        "rows": rows,
    }
    write_json(OUTPUT, payload)
    print(json.dumps({"targets": len(rows), "pages": sum(page_counts.values()), "worksheets": len(sheet_records)}, sort_keys=True))


if __name__ == "__main__":
    main()
