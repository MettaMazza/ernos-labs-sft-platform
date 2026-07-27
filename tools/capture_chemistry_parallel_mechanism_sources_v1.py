#!/usr/bin/env python3
"""Capture the complete predeclared KIN-008 source surface after the prefetch seal."""

from __future__ import annotations

import hashlib
import json
from pathlib import Path
from urllib.request import Request, urlopen

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SPEC_PATH = ROOT / "experiments/external_sources/chemistry/parallel_mechanism_capture_spec_v1.json"
SPEC_HASH = "sha256:f32b98d3cc4f02c02f01249b0f92ce799d1453ae04d1f8c9c107be6a509a6e89"
SNAPSHOT_ROOT = ROOT / "experiments/external_sources/chemistry/snapshots/kin-008-parallel-mechanism-v1"
URLS = {
    "article_html": "https://www.nature.com/articles/s41467-026-70199-4",
    "article_pdf": "https://www.nature.com/articles/s41467-026-70199-4.pdf",
    "supplementary_information": "https://static-content.springer.com/esm/art%3A10.1038%2Fs41467-026-70199-4/MediaObjects/41467_2026_70199_MOESM1_ESM.pdf",
    "transparent_peer_review": "https://static-content.springer.com/esm/art%3A10.1038%2Fs41467-026-70199-4/MediaObjects/41467_2026_70199_MOESM2_ESM.pdf",
    "source_data": "https://static-content.springer.com/esm/art%3A10.1038%2Fs41467-026-70199-4/MediaObjects/41467_2026_70199_MOESM3_ESM.xlsx",
}
FILE_NAMES = {
    "article_html": "article.html",
    "article_pdf": "article.pdf",
    "supplementary_information": "supplementary-information.pdf",
    "transparent_peer_review": "transparent-peer-review.pdf",
    "source_data": "source-data.xlsx",
}


def sha_bytes(value: bytes) -> str:
    return "sha256:" + hashlib.sha256(value).hexdigest()


def sha_file(path: Path) -> str:
    return sha_bytes(path.read_bytes())


def fetch(url: str) -> bytes:
    request = Request(url, headers={"User-Agent": "Ernos-Labs-SFT-v3-source-capture/1"})
    with urlopen(request, timeout=180) as response:
        return response.read()


def main() -> None:
    if sha_file(SPEC_PATH) != SPEC_HASH:
        raise ValueError("KIN-008 prefetch specification changed")
    spec = json.loads(SPEC_PATH.read_text())
    if (
        spec.get("target_values_or_hashes_present") is not False
        or spec.get("selection_fit_average_interpolation_renormalization_inference_or_correction_permitted") is not False
    ):
        raise ValueError("KIN-008 prefetch boundary changed")
    SNAPSHOT_ROOT.mkdir(parents=True, exist_ok=True)
    records = []
    for source_class, url in URLS.items():
        path = SNAPSHOT_ROOT / FILE_NAMES[source_class]
        content = path.read_bytes() if path.exists() else fetch(url)
        path.write_bytes(content)
        records.append({
            "source_class": source_class,
            "url": url,
            "snapshot_path": str(path.relative_to(ROOT)),
            "snapshot_hash": sha_bytes(content),
            "byte_count": len(content),
        })
    workbook_path = SNAPSHOT_ROOT / FILE_NAMES["source_data"]
    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    workbook_topology = tuple({
        "source_sheet_ordinal": ordinal,
        "source_sheet_identity": sheet.title,
        "declared_max_row": sheet.max_row,
        "declared_max_column": sheet.max_column,
    } for ordinal, sheet in enumerate(workbook.worksheets, start=1))
    if not workbook_topology:
        raise ValueError("KIN-008 source-data workbook has no worksheet")
    inventory = {
        "schema": "sft-v3-parallel-mechanism-source-inventory/1",
        "prefetch_spec_hash": SPEC_HASH,
        "article_doi": "10.1038/s41467-026-70199-4",
        "complete_source_file_count": len(records),
        "complete_source_files": records,
        "workbook_topology_only": workbook_topology,
        "worksheet_cell_values_or_hashes_present": False,
        "all_complete_rectangular_worksheet_surfaces_registered_without_value_selection": True,
    }
    inventory_path = SNAPSHOT_ROOT / "source-inventory-v1.json"
    inventory_path.write_text(json.dumps(inventory, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    print(json.dumps({
        "inventory_path": str(inventory_path.relative_to(ROOT)),
        "inventory_hash": sha_file(inventory_path),
        "complete_source_file_count": len(records),
        "workbook_topology": workbook_topology,
    }, indent=2))


if __name__ == "__main__":
    main()
